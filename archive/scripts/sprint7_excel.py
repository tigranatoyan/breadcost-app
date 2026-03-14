#!/usr/bin/env python3
"""
sprint7_excel.py — Generates Excel report with time each Sprint 7 ticket spent at each stage.
Pulls actual changelog data from Jira for accurate timestamps.
"""
import base64, urllib.request, urllib.error, json, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN

def jira_req(path):
    token = base64.b64encode((JIRA_EMAIL + ":" + JIRA_API_TOKEN).encode()).decode()
    hdrs = {
        "Authorization": "Basic " + token,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    req = urllib.request.Request(JIRA_BASE_URL + path, headers=hdrs, method="GET")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def parse_jira_date(s):
    """Parse Jira datetime string like '2025-01-15T10:30:00.000+0000'."""
    if not s:
        return None
    # Handle both +0000 and +00:00 formats
    s = s.replace("+0000", "+00:00").replace("Z", "+00:00")
    # Strip to seconds precision
    if "." in s:
        base, rest = s.split(".", 1)
        tz = ""
        for c in ["+", "-"]:
            if c in rest:
                idx = rest.index(c)
                tz = rest[idx:]
                break
        s = base + tz
    try:
        return datetime.datetime.fromisoformat(s)
    except ValueError:
        return None


STORY_KEYS = ["BC-241", "BC-242", "BC-243", "BC-244", "BC-227", "BC-228", "BC-229"]

print("Fetching changelog data from Jira...")
rows = []

for key in STORY_KEYS:
    issue = jira_req(f"/rest/api/3/issue/{key}?expand=changelog")
    summary = issue["fields"]["summary"]
    created = issue["fields"]["created"]

    # Extract status transitions from changelog
    transitions = []
    transitions.append(("Created", "To Do", parse_jira_date(created)))

    for history in issue.get("changelog", {}).get("histories", []):
        ts = parse_jira_date(history["created"])
        for item in history["items"]:
            if item["field"] == "status":
                transitions.append((item["fromString"], item["toString"], ts))

    # Calculate time at each stage
    stage_times = {}
    for i, (from_status, to_status, ts) in enumerate(transitions):
        if i == 0:
            stage_times["To Do"] = {"entered": ts, "exited": None, "duration": None}
        else:
            # Close the previous stage
            prev_stage = from_status
            if prev_stage in stage_times and stage_times[prev_stage]["exited"] is None:
                stage_times[prev_stage]["exited"] = ts
                stage_times[prev_stage]["duration"] = ts - stage_times[prev_stage]["entered"]
            # Open the new stage
            if to_status not in stage_times:
                stage_times[to_status] = {"entered": ts, "exited": None, "duration": None}

    row = {
        "key": key,
        "summary": summary,
        "created": parse_jira_date(created),
        "current_status": issue["fields"]["status"]["name"],
        "to_do_entered": stage_times.get("To Do", {}).get("entered"),
        "to_do_duration": stage_times.get("To Do", {}).get("duration"),
        "in_progress_entered": stage_times.get("In Progress", {}).get("entered"),
        "in_progress_duration": stage_times.get("In Progress", {}).get("duration"),
        "done_entered": stage_times.get("Done", {}).get("entered"),
    }
    rows.append(row)
    print(f"  {key}: {len(transitions)} transitions")

# Generate Excel
print("\nGenerating Excel report...")
wb = Workbook()
ws = wb.active
ws.title = "Sprint 7 Time Tracking"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
date_fmt = "YYYY-MM-DD HH:MM:SS"

# Headers
headers = [
    "Jira Key", "Summary", "Created", "Current Status",
    "To Do — Entered", "Time in To Do",
    "In Progress — Entered", "Time in In Progress",
    "Done — Entered", "Total Cycle Time"
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Data rows
for r_idx, row in enumerate(rows, 2):
    total_cycle = None
    if row["done_entered"] and row["to_do_entered"]:
        total_cycle = row["done_entered"] - row["to_do_entered"]

    values = [
        row["key"],
        row["summary"],
        row["created"].strftime("%Y-%m-%d %H:%M:%S") if row["created"] else "",
        row["current_status"],
        row["to_do_entered"].strftime("%Y-%m-%d %H:%M:%S") if row["to_do_entered"] else "",
        str(row["to_do_duration"]) if row["to_do_duration"] else "",
        row["in_progress_entered"].strftime("%Y-%m-%d %H:%M:%S") if row["in_progress_entered"] else "",
        str(row["in_progress_duration"]) if row["in_progress_duration"] else "",
        row["done_entered"].strftime("%Y-%m-%d %H:%M:%S") if row["done_entered"] else "",
        str(total_cycle) if total_cycle else "",
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=r_idx, column=col, value=val)
        cell.border = thin_border
        if col in (1, 4):
            cell.alignment = Alignment(horizontal="center")

# Column widths
col_widths = [12, 65, 20, 14, 20, 18, 20, 18, 20, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

# Summary section
summary_row = len(rows) + 3
ws.cell(row=summary_row, column=1, value="Sprint 7 Summary").font = Font(bold=True, size=12)
ws.cell(row=summary_row + 1, column=1, value="Sprint Name:")
ws.cell(row=summary_row + 1, column=2, value="R1.5 Sprint 7 — Reports, Dashboard & Production Polish FE")
ws.cell(row=summary_row + 2, column=1, value="Total Stories:")
ws.cell(row=summary_row + 2, column=2, value=len(rows))
ws.cell(row=summary_row + 3, column=1, value="Status:")
ws.cell(row=summary_row + 3, column=2, value="All Done — Sprint Closed")
ws.cell(row=summary_row + 4, column=1, value="Generated:")
ws.cell(row=summary_row + 4, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

# Freeze panes
ws.freeze_panes = "A2"

# Save
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "sprint7_time_report.xlsx")
output_path = os.path.normpath(output_path)
wb.save(output_path)
print(f"\nExcel report saved to: {output_path}")
print("Done!")
