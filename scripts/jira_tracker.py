"""
Jira ticket phase tracker — records timestamps when tickets transition between states.
Produces an XLSX report of time spent in each phase.

Usage:
  python jira_tracker.py transition BC-302 in_progress   # Move to In Progress
  python jira_tracker.py transition BC-302 done           # Move to Done
  python jira_tracker.py status                           # Show current sprint status
  python jira_tracker.py report                           # Generate XLSX report
"""
import urllib.request, urllib.error, base64, json, sys, os, csv
from datetime import datetime

sys.path.insert(0, ".")
from config import JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN, JIRA_PROJECT

creds = base64.b64encode((JIRA_EMAIL + ":" + JIRA_API_TOKEN).encode()).decode()
headers_get = {"Authorization": "Basic " + creds, "Accept": "application/json"}
headers_post = {
    "Authorization": "Basic " + creds,
    "Accept": "application/json",
    "Content-Type": "application/json",
}

# Transition IDs from Jira workflow
TRANSITIONS = {
    "to_do": "11",
    "in_progress": "21",
    "done": "31",
}

# Local tracking file (CSV)
TRACKING_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "jira_transitions.csv")
os.makedirs(os.path.dirname(TRACKING_FILE), exist_ok=True)

def api_get(path):
    req = urllib.request.Request(JIRA_BASE_URL + path, headers=headers_get)
    return json.loads(urllib.request.urlopen(req, timeout=15).read())


def api_post(path, body):
    data = json.dumps(body).encode()
    req = urllib.request.Request(JIRA_BASE_URL + path, data=data, headers=headers_post, method="POST")
    try:
        return json.loads(urllib.request.urlopen(req, timeout=15).read())
    except urllib.error.HTTPError as e:
        # transition POST returns 204 with no body
        if e.code == 204:
            return {}
        raise


def log_transition(ticket_key, from_state, to_state):
    """Append a row to the local CSV tracking file."""
    now = datetime.now().isoformat()
    row = [now, ticket_key, from_state, to_state]
    file_exists = os.path.exists(TRACKING_FILE)
    with open(TRACKING_FILE, "a", newline="") as f:
        w = csv.writer(f)
        if not file_exists:
            w.writerow(["timestamp", "ticket", "from_state", "to_state"])
        w.writerow(row)
    return now


def do_transition(ticket_key, target_state):
    """Transition a Jira ticket and log it locally."""
    target_state = target_state.lower().replace("-", "_").replace(" ", "_")
    if target_state not in TRANSITIONS:
        print("ERROR: Unknown state '%s'. Use: %s" % (target_state, ", ".join(TRANSITIONS.keys())))
        sys.exit(1)

    # Get current status
    issue = api_get("/rest/api/3/issue/" + ticket_key + "?fields=status,summary")
    current_status = issue["fields"]["status"]["name"]
    summary = issue["fields"]["summary"][:60]

    # Map target to display name
    state_names = {"to_do": "To Do", "in_progress": "In Progress", "done": "Done"}
    target_name = state_names[target_state]

    if current_status == target_name:
        print("  %s is already '%s' — skipping" % (ticket_key, target_name))
        return

    # Execute transition
    trans_id = TRANSITIONS[target_state]
    body = {"transition": {"id": trans_id}}
    try:
        api_post("/rest/api/3/issue/" + ticket_key + "/transitions", body)
    except urllib.error.HTTPError as e:
        if e.code != 204:
            print("ERROR transitioning %s: %s" % (ticket_key, e.read().decode()[:200]))
            sys.exit(1)

    # Log locally
    ts = log_transition(ticket_key, current_status, target_name)
    print("  %s  %s: %s -> %s" % (ts[:19], ticket_key, current_status, target_name))
    print("  %s" % summary)


def show_status():
    """Show current sprint tickets and their states."""
    # Get active sprint
    sprints = api_get("/rest/agile/1.0/board/2/sprint?state=active")
    if not sprints.get("values"):
        print("No active sprint")
        return

    sprint = sprints["values"][0]
    print("Sprint: %s (id=%s)" % (sprint["name"], sprint["id"]))
    print("  Start: %s  End: %s" % (sprint.get("startDate", "?")[:10], sprint.get("endDate", "?")[:10]))
    print()

    # Get sprint issues
    issues = api_get("/rest/agile/1.0/sprint/%s/issue?maxResults=50" % sprint["id"])
    by_status = {"To Do": [], "In Progress": [], "Done": []}
    for iss in issues.get("issues", []):
        status = iss["fields"]["status"]["name"]
        key = iss["key"]
        priority = iss["fields"].get("priority", {}).get("name", "?")
        summary = iss["fields"]["summary"][:55]
        entry = "%s  %-8s  %s" % (key, priority, summary)
        if status in by_status:
            by_status[status].append(entry)
        else:
            by_status.setdefault(status, []).append(entry)

    for state in ["To Do", "In Progress", "Done"]:
        items = by_status.get(state, [])
        print("--- %s (%d) ---" % (state, len(items)))
        for item in items:
            print("  " + item)
        print()


def generate_report():
    """Generate XLSX report from local transition log."""
    if not os.path.exists(TRACKING_FILE):
        print("No transition data yet — use 'transition' command first.")
        return

    # Read all transitions
    with open(TRACKING_FILE, "r") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        print("No transitions recorded yet.")
        return

    # Build per-ticket timeline
    tickets = {}
    for row in rows:
        key = row["ticket"]
        if key not in tickets:
            tickets[key] = []
        tickets[key].append({
            "timestamp": datetime.fromisoformat(row["timestamp"]),
            "from": row["from_state"],
            "to": row["to_state"],
        })

    # Calculate time in each state
    report_rows = []
    for key, transitions in sorted(tickets.items()):
        transitions.sort(key=lambda x: x["timestamp"])
        time_in_progress = 0.0
        time_to_done = 0.0

        first_in_progress = None
        last_in_progress_start = None
        done_time = None

        for t in transitions:
            if t["to"] == "In Progress" and last_in_progress_start is None:
                last_in_progress_start = t["timestamp"]
                if first_in_progress is None:
                    first_in_progress = t["timestamp"]
            elif t["to"] in ("Done", "To Do") and last_in_progress_start is not None:
                delta = (t["timestamp"] - last_in_progress_start).total_seconds() / 3600
                time_in_progress += delta
                last_in_progress_start = None
            if t["to"] == "Done":
                done_time = t["timestamp"]

        if first_in_progress and done_time:
            time_to_done = (done_time - first_in_progress).total_seconds() / 3600

        started = transitions[0]["timestamp"].strftime("%Y-%m-%d %H:%M")
        ended = done_time.strftime("%Y-%m-%d %H:%M") if done_time else "—"
        status = transitions[-1]["to"]

        report_rows.append({
            "Ticket": key,
            "Status": status,
            "Started": started,
            "Completed": ended,
            "Hours In Progress": round(time_in_progress, 2),
            "Hours Total (Start to Done)": round(time_to_done, 2) if time_to_done else "—",
        })

    # Try openpyxl for XLSX, fall back to CSV
    report_path = os.path.join(os.path.dirname(TRACKING_FILE), "sprint_report.xlsx")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Sprint Time Report"

        # Header
        header_cols = ["Ticket", "Status", "Started", "Completed", "Hours In Progress", "Hours Total (Start to Done)"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(header_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        status_colors = {
            "Done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "In Progress": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "To Do": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

        for row_idx, row_data in enumerate(report_rows, 2):
            for col_idx, col_name in enumerate(header_cols, 1):
                val = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_name == "Status":
                    cell.fill = status_colors.get(val, PatternFill())

        # Column widths
        col_widths = [12, 14, 18, 18, 18, 28]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        wb.save(report_path)
        print("XLSX report saved: %s" % report_path)
        print("  %d tickets tracked" % len(report_rows))

    except ImportError:
        # Fallback to CSV
        report_path = report_path.replace(".xlsx", ".csv")
        with open(report_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["Ticket", "Status", "Started", "Completed", "Hours In Progress", "Hours Total (Start to Done)"])
            w.writeheader()
            w.writerows(report_rows)
        print("CSV report saved (install openpyxl for XLSX): %s" % report_path)
        print("  %d tickets tracked" % len(report_rows))


# ────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    cmd = sys.argv[1].lower()

    if cmd == "transition":
        if len(sys.argv) < 4:
            print("Usage: python jira_tracker.py transition <TICKET> <STATE>")
            print("States: to_do, in_progress, done")
            sys.exit(1)
        do_transition(sys.argv[2].upper(), sys.argv[3])

    elif cmd == "status":
        show_status()

    elif cmd == "report":
        generate_report()

    else:
        print("Unknown command: %s" % cmd)
        print("Commands: transition, status, report")
        sys.exit(1)
